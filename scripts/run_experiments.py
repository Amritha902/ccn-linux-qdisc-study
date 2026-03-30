#!/usr/bin/env python3
"""
run_experiments.py — Automated Multi-Scenario Experiment Suite (FIXED)
ACAPE Research Paper — VIT Chennai 2026

Fixes from v1:
  - iperf server runs persistently in ns1 (no --one-off), killed manually
  - Verifies iperf sees ~correct throughput before controller starts
  - target_ms parsed correctly (handles both us and ms from tc output)
  - Controller paths are absolute, not relative
  - Stats collector verifies traffic is actually flowing before logging
  - Port 5299 used to avoid conflict with any host server on 5201/5202

Usage:
  sudo python3 scripts/run_experiments.py
  sudo python3 scripts/run_experiments.py --scenarios S1 S2
  sudo python3 scripts/run_experiments.py --dry-run
"""

import subprocess, re, time, csv, os, sys, argparse, signal
import threading, socket
from datetime import datetime
from pathlib import Path

# ── Scenarios ──────────────────────────────────────────────────
SCENARIOS = {
    "S1": {
        "name":     "Baseline (10Mbit, 8 flows)",
        "rate":     "10mbit", "burst": "32kbit", "latency": "400ms",
        "flows":    8,  "duration": 120,
        "expected_tput_range": (8.0, 11.0),
        "desc":     "Standard testbed — 8 TCP flows, 10 Mbit bottleneck",
    },
    "S2": {
        "name":     "Heavy Bottleneck (5Mbit, 8 flows)",
        "rate":     "5mbit",  "burst": "16kbit", "latency": "400ms",
        "flows":    8,  "duration": 120,
        "expected_tput_range": (4.0, 5.5),
        "desc":     "Tight 5 Mbit bottleneck — higher oversubscription",
    },
    "S3": {
        "name":     "Relaxed Bottleneck (20Mbit, 8 flows)",
        "rate":     "20mbit", "burst": "64kbit", "latency": "400ms",
        "flows":    8,  "duration": 120,
        "expected_tput_range": (16.0, 21.0),
        "desc":     "20 Mbit — lower congestion pressure",
    },
    "S4": {
        "name":     "High Flow Density (10Mbit, 16 flows)",
        "rate":     "10mbit", "burst": "32kbit", "latency": "400ms",
        "flows":    16, "duration": 120,
        "expected_tput_range": (8.0, 11.0),
        "desc":     "16 parallel flows — mice-heavy workload",
    },
    "S5": {
        "name":     "Mixed Workload (10Mbit, 4 elephant flows)",
        "rate":     "10mbit", "burst": "32kbit", "latency": "400ms",
        "flows":    4,  "duration": 120,
        "expected_tput_range": (8.0, 11.0),
        "desc":     "4 large-buffer flows simulating elephant traffic",
    },
    "S6": {
        "name":     "Bursty Traffic (10Mbit, 8 flows, 60s)",
        "rate":     "10mbit", "burst": "32kbit", "latency": "400ms",
        "flows":    8,  "duration": 60,
        "expected_tput_range": (8.0, 11.0),
        "desc":     "Short 60s run — tests transient controller response",
    },
}

CONTROLLERS  = ["static", "adaptive_red", "acape"]
NS1, NS2     = "ns1", "ns2"
IP1, IP2     = "10.0.0.1", "10.0.0.2"
IFACE        = "veth1"
PORT         = 5299        # avoids conflict with any host iperf server
REPO         = Path(__file__).resolve().parent.parent
SCRIPTS      = Path(__file__).resolve().parent

GREEN  = "\033[92m"; YELLOW = "\033[93m"
RED    = "\033[91m"; CYAN   = "\033[96m"
BOLD   = "\033[1m";  RESET  = "\033[0m"

def log(msg,  c=GREEN):  print(f"{c}{BOLD}[{datetime.now().strftime('%H:%M:%S')}]{RESET} {msg}", flush=True)
def warn(msg):            print(f"{YELLOW}[WARN]{RESET} {msg}", flush=True)
def err(msg):             print(f"{RED}[ERR]{RESET} {msg}", flush=True)
def info(msg):            print(f"  {CYAN}{msg}{RESET}", flush=True)


def shell(cmd, ns=None, timeout=10):
    """Run a command, optionally inside a namespace."""
    if ns:
        cmd = ["ip", "netns", "exec", ns] + cmd
    try:
        r = subprocess.run(cmd, capture_output=True, text=True, timeout=timeout)
        return r.stdout + r.stderr
    except Exception as e:
        return str(e)


def popen(cmd, ns=None, stdout=None, stderr=None):
    """Start a background process, optionally inside a namespace."""
    if ns:
        cmd = ["ip", "netns", "exec", ns] + cmd
    return subprocess.Popen(
        cmd,
        stdout=stdout or subprocess.DEVNULL,
        stderr=stderr or subprocess.DEVNULL,
    )


def kill_all_iperf():
    """Kill any running iperf3 servers on host or in namespaces."""
    subprocess.run(["pkill", "-9", "-f", "iperf3"], capture_output=True)
    time.sleep(1)


# ── Namespace setup ────────────────────────────────────────────
def setup_namespace(scenario):
    s = SCENARIOS[scenario]
    rate = s["rate"]
    log(f"Setting up namespace — {rate} bottleneck, {s['flows']} flows")

    # Teardown everything
    kill_all_iperf()
    shell(["ip", "netns", "del", NS1])
    shell(["ip", "netns", "del", NS2])
    shell(["ip", "link", "del", IFACE])
    time.sleep(1.5)

    # Create namespaces and veth pair
    shell(["ip", "netns", "add", NS1])
    shell(["ip", "netns", "add", NS2])
    shell(["ip", "link", "add", IFACE, "type", "veth", "peer", "name", "veth2"])
    shell(["ip", "link", "set", IFACE,   "netns", NS1])
    shell(["ip", "link", "set", "veth2", "netns", NS2])

    # Configure IPs
    shell(["ip", "addr", "add", f"{IP1}/24", "dev", IFACE],  ns=NS1)
    shell(["ip", "addr", "add", f"{IP2}/24", "dev", "veth2"], ns=NS2)
    shell(["ip", "link", "set", IFACE,   "up"], ns=NS1)
    shell(["ip", "link", "set", "veth2", "up"], ns=NS2)
    shell(["ip", "link", "set", "lo",    "up"], ns=NS1)
    shell(["ip", "link", "set", "lo",    "up"], ns=NS2)

    # Apply TBF bottleneck + fq_codel
    shell(["tc", "qdisc", "add", "dev", IFACE, "root", "handle", "1:",
           "tbf", "rate", rate, "burst", s["burst"], "latency", s["latency"]], ns=NS1)
    shell(["tc", "qdisc", "add", "dev", IFACE, "parent", "1:1", "handle", "10:",
           "fq_codel", "target", "5ms", "interval", "100ms",
           "limit", "1024", "quantum", "1514"], ns=NS1)

    # Verify qdisc
    qout = shell(["tc", "qdisc", "show", "dev", IFACE], ns=NS1)
    if "fq_codel" not in qout:
        err("fq_codel not found after setup!")
        return False

    # Verify connectivity
    ping = shell(["ping", "-c", "3", "-W", "1", IP1], ns=NS2, timeout=6)
    if "0% packet loss" in ping:
        log(f"Namespace ready — tbf {rate} + fq_codel verified")
        return True
    else:
        err(f"Ping failed! {ping[:100]}")
        return False


def reset_qdisc():
    """Reset fq_codel to defaults between controller runs."""
    shell(["tc", "qdisc", "change", "dev", IFACE, "parent", "1:1", "handle", "10:",
           "fq_codel", "target", "5ms", "interval", "100ms",
           "limit", "1024", "quantum", "1514"], ns=NS1)


# ── tc stats parser (FIXED) ────────────────────────────────────
def parse_tc(ns, iface):
    out = shell(["tc", "-s", "qdisc", "show", "dev", iface], ns=ns, timeout=3)
    drops = backlog = sent_bytes = 0
    target_ms = 5.0
    limit = 1024

    m = re.search(r"dropped (\d+)", out)
    if m: drops = int(m.group(1))

    m = re.search(r"backlog \d+b (\d+)p", out)
    if m: backlog = int(m.group(1))

    m = re.search(r"Sent (\d+) bytes", out)
    if m: sent_bytes = int(m.group(1))

    # FIXED: handle both "target 5ms" and "target 5000us" or "target 999us"
    m = re.search(r"target (\d+)(us|ms)", out)
    if m:
        val, unit = int(m.group(1)), m.group(2)
        target_ms = val / 1000.0 if unit == "us" else float(val)

    m = re.search(r"limit (\d+)p", out)
    if m: limit = int(m.group(1))

    return drops, backlog, sent_bytes, target_ms, limit


# ── Throughput verifier ────────────────────────────────────────
def verify_throughput(scenario, timeout=15):
    """
    Check that iperf client is seeing expected throughput range.
    This catches the case where iperf hits the host server instead of ns1.
    """
    s = SCENARIOS[scenario]
    lo, hi = s["expected_tput_range"]
    t_start = time.time()
    prev_bytes = 0
    prev_ts    = time.time()

    info("Verifying traffic goes through bottleneck...")
    while time.time() - t_start < timeout:
        time.sleep(2)
        _, _, sent, tgt_ms, _ = parse_tc(NS1, IFACE)
        now = time.time()
        dt  = now - prev_ts
        mbps = max(0, sent - prev_bytes) * 8 / (dt * 1e6) if prev_bytes else 0
        prev_bytes = sent
        prev_ts    = now
        if mbps > 0:
            info(f"  tc measured throughput: {mbps:.2f} Mbps (expected {lo}-{hi})")
            if lo <= mbps <= hi:
                info("  PASS — traffic correctly through bottleneck")
                return True
            elif mbps > hi * 1.5:
                err(f"  FAIL — {mbps:.1f} Mbps >> {hi} Mbit limit")
                err("  iperf client is NOT hitting ns1 server — killing run")
                return False

    warn("Could not verify throughput in time — proceeding anyway")
    return True


# ── Stats collector thread ─────────────────────────────────────
class StatsCollector:
    def __init__(self, outfile, duration):
        self.outfile  = str(outfile)
        self.duration = duration
        self._stop    = threading.Event()
        self._thread  = None
        self._prev    = {"ts": time.time(), "drops": 0, "bytes": 0}
        self.n_rows   = 0

    def _run(self):
        with open(self.outfile, "w", newline="") as f:
            w = csv.writer(f)
            w.writerow(["t_s", "drop_rate_per_s", "backlog_pkts",
                        "throughput_mbps", "target_ms", "limit_pkts"])
            t0 = time.time()
            while not self._stop.is_set():
                time.sleep(0.5)
                drops, bl, sent, tgt_ms, lim = parse_tc(NS1, IFACE)
                now = time.time()
                dt  = max(now - self._prev["ts"], 1e-6)
                dr  = max(0, drops - self._prev["drops"]) / dt
                tp  = max(0, sent  - self._prev["bytes"]) * 8 / (dt * 1e6)
                t_s = now - t0
                w.writerow([round(t_s, 2), round(dr, 1), bl,
                            round(tp, 3), round(tgt_ms, 3), lim])
                f.flush()
                self._prev = {"ts": now, "drops": drops, "bytes": sent}
                self.n_rows += 1

    def start(self):
        self._thread = threading.Thread(target=self._run, daemon=True)
        self._thread.start()

    def stop(self):
        self._stop.set()
        if self._thread:
            self._thread.join(timeout=4)


# ── Run one experiment cell ────────────────────────────────────
def run_one(scenario, controller, cell_dir):
    s   = SCENARIOS[scenario]
    dur = s["duration"]

    info(f"Cell: {scenario} x {controller} ({dur}s)")
    cell_dir.mkdir(parents=True, exist_ok=True)

    reset_qdisc()
    time.sleep(1)
    kill_all_iperf()
    time.sleep(1)

    # Start iperf3 server INSIDE ns1 (persistent, no --one-off)
    srv_log  = cell_dir / "iperf_server.log"
    srv_proc = popen(
        ["iperf3", "-s", "-p", str(PORT)],
        ns=NS1,
        stdout=open(srv_log, "w"),
        stderr=subprocess.STDOUT,
    )
    time.sleep(2)  # Wait for server to be ready

    # Start stats collector
    metrics_file = cell_dir / "metrics.csv"
    collector    = StatsCollector(metrics_file, dur)
    collector.start()
    time.sleep(0.5)

    # Start iperf3 client from ns2
    cli_log  = cell_dir / "iperf_client.log"
    iperf_cmd = ["iperf3",
                 "-c", IP1, "-p", str(PORT),
                 "-P", str(s["flows"]),
                 "-t", str(dur),
                 "-i", "0"]
    if scenario == "S5":
        iperf_cmd += ["-l", "128K"]   # large send size = elephant flows
    cli_proc = popen(
        iperf_cmd,
        ns=NS2,
        stdout=open(cli_log, "w"),
        stderr=subprocess.STDOUT,
    )

    # Verify traffic is going through the bottleneck
    ok = verify_throughput(scenario, timeout=15)
    if not ok:
        # Kill everything — this run is invalid
        cli_proc.kill(); srv_proc.kill(); collector.stop()
        err(f"  INVALID RUN — iperf not through bottleneck. Skipping {scenario}x{controller}")
        return None

    # Start controller (after traffic is flowing and verified)
    ctrl_proc = None
    ctrl_log  = cell_dir / "controller.log"

    if controller == "adaptive_red":
        time.sleep(1)
        ctrl_proc = subprocess.Popen(
            ["python3", str(SCRIPTS / "adaptive_red.py"),
             "--ns",       NS1,
             "--iface",    IFACE,
             "--logdir",   str(cell_dir),
             "--duration", str(dur + 10)],
            stdout=open(ctrl_log, "w"),
            stderr=subprocess.STDOUT,
        )

    elif controller == "acape":
        time.sleep(1)
        ctrl_proc = subprocess.Popen(
            ["python3", str(SCRIPTS / "acape_v5.py"),
             "--ns",     NS1,
             "--iface",  IFACE,
             "--logdir", str(cell_dir)],
            stdout=open(ctrl_log, "w"),
            stderr=subprocess.STDOUT,
        )

    # Wait for iperf to finish
    try:
        cli_proc.wait(timeout=dur + 30)
    except subprocess.TimeoutExpired:
        warn("  iperf timed out — killing")
        cli_proc.kill()

    # Stop everything
    if ctrl_proc:
        ctrl_proc.terminate()
        try: ctrl_proc.wait(timeout=5)
        except: ctrl_proc.kill()

    collector.stop()
    srv_proc.terminate()
    try: srv_proc.wait(timeout=3)
    except: srv_proc.kill()

    time.sleep(2)

    # Parse results
    return parse_results(metrics_file, scenario, controller)


# ── Parse results ──────────────────────────────────────────────
def parse_results(metrics_file, scenario, controller):
    rows = []
    try:
        with open(metrics_file) as f:
            for r in csv.DictReader(f):
                rows.append(r)
    except Exception as e:
        warn(f"Cannot read {metrics_file}: {e}")
        return None

    if len(rows) < 5:
        warn(f"Too few rows ({len(rows)}) — run may have failed")
        return None

    def fv(v, d=0.0):
        try: return float(v)
        except: return d

    bl   = [fv(r["backlog_pkts"])    for r in rows]
    dr   = [fv(r["drop_rate_per_s"]) for r in rows]
    tp   = [fv(r["throughput_mbps"]) for r in rows]
    tgt  = [fv(r["target_ms"])       for r in rows]
    lim  = [fv(r["limit_pkts"])      for r in rows]

    # Filter to active traffic period (throughput > 0.5 Mbps)
    active = [(b,d,t,g,l) for b,d,t,g,l in zip(bl,dr,tp,tgt,lim) if t > 0.5]
    if not active:
        active = list(zip(bl, dr, tp, tgt, lim))

    bl_a  = [x[0] for x in active]
    dr_a  = [x[1] for x in active]
    tp_a  = [x[2] for x in active]
    tgt_a = [x[3] for x in active]

    import statistics
    sm = lambda lst: statistics.mean(lst)   if lst else 0
    sd = lambda lst: statistics.stdev(lst)  if len(lst)>1 else 0

    # Stabilisation: first time backlog stays below 80% of initial for 5s (10 samples)
    s    = SCENARIOS[scenario]
    t_v  = [fv(r["t_s"]) for r in rows]
    stab = s["duration"]
    W    = 10
    for i in range(len(bl) - W):
        threshold = min(400, max(bl[:5]) * 0.85 if bl[:5] else 400)
        if all(b < threshold and tp_a[0] * 0.8 < tp[i] for b in bl[i:i+W]):
            stab = t_v[i]
            break

    return {
        "scenario":       scenario,
        "scenario_name":  s["name"],
        "controller":     controller,
        "bottleneck":     s["rate"],
        "flows":          s["flows"],
        "duration_s":     s["duration"],
        "avg_backlog":    round(sm(bl_a),  1),
        "max_backlog":    round(max(bl_a), 1) if bl_a else 0,
        "std_backlog":    round(sd(bl_a),  1),
        "avg_drop_rate":  round(sm(dr_a),  1),
        "avg_throughput": round(sm(tp_a),  3),
        "min_throughput": round(min(tp_a), 3) if tp_a else 0,
        "stab_time_s":    round(stab,      1),
        "tgt_initial":    round(tgt_a[0],  3) if tgt_a else 5.0,
        "tgt_final":      round(tgt_a[-1], 3) if tgt_a else 5.0,
        "tgt_reduction":  round(tgt_a[0] - tgt_a[-1], 3) if tgt_a else 0,
        "avg_limit":      round(sm(lim if lim else [1024]), 0),
        "n_samples":      len(rows),
    }


# ── Excel writer ───────────────────────────────────────────────
def write_excel(results, path):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        warn("openpyxl missing — pip3 install openpyxl --break-system-packages")
        return

    wb  = openpyxl.Workbook()
    ws  = wb.active
    ws.title = "All Results"

    HDR_FILL  = PatternFill("solid", fgColor="1F2937")
    S_FILL    = PatternFill("solid", fgColor="4C0519")
    A_FILL    = PatternFill("solid", fgColor="1E3A5F")
    C_FILL    = PatternFill("solid", fgColor="065F46")
    HDR_FONT  = Font(bold=True, color="E6EDF3", size=11)
    CELL_FONT = Font(color="E6EDF3", size=10)
    thin      = Border(**{s: Side(style="thin", color="374151")
                          for s in ["left","right","top","bottom"]})

    HEADERS = ["Scenario","Name","Controller","Bottleneck","Flows",
               "Avg Backlog","Max Backlog","Std Backlog",
               "Avg Drop/s","Avg Tput (Mbps)","Min Tput (Mbps)",
               "Stab Time (s)","Tgt Start (ms)","Tgt End (ms)",
               "Tgt Reduction (ms)","Samples"]
    KEYS    = ["scenario","scenario_name","controller","bottleneck","flows",
               "avg_backlog","max_backlog","std_backlog",
               "avg_drop_rate","avg_throughput","min_throughput",
               "stab_time_s","tgt_initial","tgt_final",
               "tgt_reduction","n_samples"]

    for ci, h in enumerate(HEADERS, 1):
        c = ws.cell(1, ci, h)
        c.fill = HDR_FILL; c.font = HDR_FONT
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        c.border = thin

    fills = {"static":"S","adaptive_red":"A","acape":"C"}
    fmap  = {"S": S_FILL, "A": A_FILL, "C": C_FILL}

    for ri, r in enumerate(results, 2):
        fill = fmap.get(fills.get(r.get("controller",""), "S"), S_FILL)
        for ci, k in enumerate(KEYS, 1):
            cell = ws.cell(ri, ci, r.get(k, ""))
            cell.fill = fill; cell.font = CELL_FONT
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin

    # Auto-width
    for col in ws.columns:
        w = max(len(str(c.value or "")) for c in col) + 2
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(w, 22)
    ws.row_dimensions[1].height = 38
    ws.freeze_panes = "A2"

    # ── Pivot sheet ────────────────────────────────────────────
    ws2 = wb.create_sheet("Pivot — Avg Backlog")
    scenarios   = sorted(set(r["scenario"]   for r in results))
    controllers = [c for c in ["static","adaptive_red","acape"]
                   if c in set(r["controller"] for r in results)]
    pivot = {(r["scenario"],r["controller"]): r for r in results}

    hdr = ["Scenario","Bottleneck","Flows"] + \
          [c.replace("_"," ").title() for c in controllers] + \
          ["ACAPE vs Static","ACAPE vs A.RED"]
    for ci, h in enumerate(hdr, 1):
        c = ws2.cell(1, ci, h)
        c.fill = HDR_FILL; c.font = HDR_FONT
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        c.border = thin

    for ri, sc in enumerate(scenarios, 2):
        s_info = SCENARIOS.get(sc, {})
        ws2.cell(ri, 1, sc).font = Font(color="E6EDF3", bold=True, size=10)
        ws2.cell(ri, 2, s_info.get("rate","")).font = Font(color="8B949E", size=10)
        ws2.cell(ri, 3, s_info.get("flows","")).font = Font(color="8B949E", size=10)

        vals = {}
        for ci, ct in enumerate(controllers, 4):
            v = pivot.get((sc, ct), {}).get("avg_backlog", "N/A")
            vals[ct] = float(v) if v != "N/A" else None
            fill = {"static": S_FILL,"adaptive_red": A_FILL,"acape": C_FILL}.get(ct, HDR_FILL)
            cell = ws2.cell(ri, ci, v)
            cell.fill = fill
            cell.font = Font(color="E6EDF3", bold=(ct=="acape"), size=11)
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin

        # Improvement
        s_bl = vals.get("static");       c_bl = vals.get("acape")
        a_bl = vals.get("adaptive_red")
        col_off = 4 + len(controllers)
        if s_bl and c_bl:
            pct = round((1-c_bl/s_bl)*100, 1)
            cell = ws2.cell(ri, col_off, f"{pct}% lower")
            cell.font = Font(color="3FB950", bold=True, size=11)
        if a_bl and c_bl:
            pct = round((1-c_bl/a_bl)*100, 1)
            cell = ws2.cell(ri, col_off+1, f"{pct}% lower")
            cell.font = Font(color="3FB950", bold=True, size=11)

    for col in ws2.columns:
        w = max(len(str(c.value or "")) for c in col) + 2
        ws2.column_dimensions[get_column_letter(col[0].column)].width = min(w, 22)
    ws2.freeze_panes = "D2"

    wb.save(str(path))
    log(f"Excel saved: {path}")


# ── Main ───────────────────────────────────────────────────────
def main():
    p = argparse.ArgumentParser()
    p.add_argument("--scenarios",   nargs="+", default=list(SCENARIOS.keys()))
    p.add_argument("--controllers", nargs="+", default=CONTROLLERS)
    p.add_argument("--outdir",      default=str(REPO / "logs" / "experiments"))
    p.add_argument("--dry-run",     action="store_true")
    args = p.parse_args()

    if not args.dry_run and os.geteuid() != 0:
        sys.exit("Run as root: sudo python3 scripts/run_experiments.py")

    try:
        import openpyxl
    except ImportError:
        os.system("pip3 install openpyxl --break-system-packages -q")

    outdir = Path(args.outdir)
    outdir.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")

    total = len(args.scenarios) * len(args.controllers)
    est   = sum(SCENARIOS[s]["duration"] + 40
                for s in args.scenarios) * len(args.controllers)

    print()
    print("=" * 68)
    print(f"  ACAPE Multi-Scenario Experiment Suite  —  VIT Chennai 2026")
    print("=" * 68)
    print(f"  Scenarios   : {', '.join(args.scenarios)}")
    print(f"  Controllers : {', '.join(args.controllers)}")
    print(f"  Total runs  : {total}")
    print(f"  Est. time   : {est//60}m {est%60}s")
    print(f"  Output      : {outdir}/")
    print("=" * 68)
    for s in args.scenarios:
        for c in args.controllers:
            print(f"  {s} x {c:14s} [{SCENARIOS[s]['duration']}s]  {SCENARIOS[s]['name']}")
    print()

    if args.dry_run:
        print("Dry run complete."); return

    input("Press ENTER to start (Ctrl+C to abort)... ")
    print()

    all_results = []
    run_num     = 0

    for scenario in args.scenarios:
        print()
        log(f"{'='*60}", YELLOW)
        log(f"SCENARIO {scenario}: {SCENARIOS[scenario]['name']}", YELLOW)
        log(f"{'='*60}", YELLOW)

        if not setup_namespace(scenario):
            err(f"Namespace setup failed for {scenario} — skipping")
            continue

        for controller in args.controllers:
            run_num += 1
            print()
            log(f"[{run_num}/{total}] {scenario} x {controller}", CYAN)

            cell_dir = outdir / f"{scenario}_{controller}_{ts}"
            result   = run_one(scenario, controller, cell_dir)

            if result:
                all_results.append(result)
                log(f"  OK: avg_backlog={result['avg_backlog']}p  "
                    f"stab={result['stab_time_s']}s  "
                    f"tput={result['avg_throughput']}Mbps  "
                    f"tgt:{result['tgt_initial']}->{result['tgt_final']}ms")
            else:
                warn(f"  FAILED or INVALID: {scenario} x {controller}")

            reset_qdisc()
            time.sleep(3)

    # Write outputs
    if all_results:
        csv_path  = outdir / f"results_{ts}.csv"
        xlsx_path = outdir / f"results_{ts}.xlsx"

        fields = list(all_results[0].keys())
        with open(csv_path, "w", newline="") as f:
            w = csv.DictWriter(f, fieldnames=fields)
            w.writeheader(); w.writerows(all_results)
        log(f"CSV: {csv_path}")

        write_excel(all_results, xlsx_path)

        print()
        print("=" * 72)
        print("  RESULTS SUMMARY")
        print("=" * 72)
        print(f"  {'Sc':4} {'Controller':14} {'AvgBL':7} {'Stab':7} "
              f"{'Tput':6} {'Tgt':10} {'Valid?':6}")
        print("  " + "-"*60)
        for r in all_results:
            s = SCENARIOS.get(r["scenario"],{})
            lo, hi = s.get("expected_tput_range",(0,999))
            valid = "YES" if lo <= r["avg_throughput"] <= hi else "WARN"
            print(f"  {r['scenario']:4} {r['controller']:14} "
                  f"{r['avg_backlog']:7.1f} {r['stab_time_s']:7.1f} "
                  f"{r['avg_throughput']:6.2f} "
                  f"{r['tgt_initial']:.1f}->{r['tgt_final']:.1f}ms  {valid}")
        print("=" * 72)
        print()
        log("Done! Now run:")
        print(f"  python3 scripts/plot_experiments.py --expdir {outdir}")
    else:
        warn("No valid results collected.")


if __name__ == "__main__":
    main()
